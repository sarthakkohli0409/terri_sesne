from fastapi import FastAPI, UploadFile, File, Form, Response
import heapq
import gc
from scipy.spatial import cKDTree
from scipy.spatial.distance import cdist
try:
    import alphashape
    from shapely.geometry import mapping as shapely_mapping
    from shapely.geometry import Polygon as ShapelyPolygon, box as shapely_box
    from shapely.ops import unary_union
    HAS_ALPHASHAPE = True
    HAS_SHAPELY = True
except ImportError:
    HAS_ALPHASHAPE = False
    HAS_SHAPELY = False
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import pandas as pd
import numpy as np
from sklearn.cluster import KMeans
from sklearn.metrics import pairwise_distances
from geopy.distance import geodesic
import io
import folium
import os
import random
import json
import zipfile
from typing import Dict, Tuple, List
from collections import defaultdict
import math
from groq import Groq
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime

app = FastAPI()
ALLOWED_ORIGINS = [
    "https://terrisense-frontend.onrender.com",
    "https://terrisense.onrender.com",
    "http://localhost:3000",
    "http://localhost:3001",
    "*",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

# ---------------------------------------------------------------------------
# Master geodata — loaded once at startup
# ---------------------------------------------------------------------------
MASTER_ZIP_DF: pd.DataFrame = pd.DataFrame()

def load_master_geodata(filename: str = "us_zips.csv"):
    global MASTER_ZIP_DF
    current_dir = os.path.dirname(os.path.abspath(__file__))
    filepath = os.path.join(current_dir, filename)
    try:
        if os.path.exists(filepath):
            MASTER_ZIP_DF = pd.read_csv(filepath, dtype=str)
            MASTER_ZIP_DF.columns = MASTER_ZIP_DF.columns.str.strip().str.lower()
            MASTER_ZIP_DF.rename(columns={
                'zip': 'zip_code', 'zipcode': 'zip_code',
                'postal': 'zip_code', 'lat': 'latitude', 'lng': 'longitude',
                'long': 'longitude',
            }, inplace=True)
            required = {'zip_code', 'latitude', 'longitude'}
            if required.issubset(MASTER_ZIP_DF.columns):
                MASTER_ZIP_DF['zip_code'] = (
                    MASTER_ZIP_DF['zip_code'].astype(str).str.strip().str.zfill(5)
                )
                MASTER_ZIP_DF['latitude'] = pd.to_numeric(
                    MASTER_ZIP_DF['latitude'], errors='coerce')
                MASTER_ZIP_DF['longitude'] = pd.to_numeric(
                    MASTER_ZIP_DF['longitude'], errors='coerce')
                MASTER_ZIP_DF.dropna(
                    subset=['latitude', 'longitude', 'zip_code'], inplace=True)
                print(f"Master geodata loaded: {len(MASTER_ZIP_DF)} rows")
            else:
                print(f"Warning: missing columns. Found: {MASTER_ZIP_DF.columns.tolist()}")
        else:
            print(f"Error: geodata file not found at {filepath}")
    except Exception as e:
        print(f"Failed to load master geodata: {e}")

load_master_geodata()


# ---------------------------------------------------------------------------
# File reading
# ---------------------------------------------------------------------------
def read_file_smartly(content: bytes, filename: str) -> pd.DataFrame:
    filename_lower = filename.lower()
    if filename_lower.endswith('.csv'):
        df = pd.read_csv(io.StringIO(content.decode('utf-8')), dtype=str)
    elif filename_lower.endswith(('.xls', '.xlsx')):
        df = pd.read_excel(io.BytesIO(content), dtype=str)
    else:
        raise ValueError("Unsupported file type. Please upload CSV or Excel.")
    df.columns = df.columns.str.strip()
    return df


# ---------------------------------------------------------------------------
# Preprocessing
# ---------------------------------------------------------------------------
def preprocess_data(df: pd.DataFrame, column_config: Dict[str, float]) -> pd.DataFrame:
    global MASTER_ZIP_DF

    zip_col = next((
        c for c in df.columns
        if c.lower().replace('_', '').replace(' ', '')
        in ['zip', 'zipcode', 'postalcode', 'rowlabels']
    ), None)
    if not zip_col:
        raise ValueError(f"Cannot find ZIP column. Columns: {df.columns.tolist()}")

    df[zip_col] = df[zip_col].astype(str).str.split('.').str[0].str.strip()
    df['clean_zip'] = df[zip_col].apply(lambda x: x.zfill(5))

    numeric_cols = []
    for col in column_config:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            numeric_cols.append(col)

    if not numeric_cols:
        raise ValueError(
            f"None of the specified columns found. "
            f"You specified: {list(column_config.keys())}. "
            f"File has: {df.columns.tolist()}"
        )

    for col in numeric_cols:
        col_max = df[col].max()
        df[f'_norm_{col}'] = df[col] / col_max if col_max > 0 else 0.0

    df['_composite'] = 0.0
    for col, pct in column_config.items():
        if col in numeric_cols:
            df['_composite'] += (float(pct) / 100.0) * df[f'_norm_{col}']

    df_agg = df.groupby('clean_zip').agg(
        zip_composite=('_composite', 'sum'),
        rep_count=('clean_zip', 'count')
    ).reset_index()

    if MASTER_ZIP_DF.empty:
        raise ValueError("Master geodata missing or failed to load.")

    merged = pd.merge(
        df_agg,
        MASTER_ZIP_DF[['zip_code', 'latitude', 'longitude']],
        left_on='clean_zip', right_on='zip_code',
        how='left'  # only keep ZIPs that appear in user's file
    )
    merged['clean_zip'] = merged['clean_zip'].fillna(merged['zip_code'])
    merged['zip_composite'] = merged['zip_composite'].fillna(0.0)
    merged['rep_count'] = merged['rep_count'].fillna(0).astype(int)
    merged['latitude'] = merged['latitude'].astype(float)
    merged['longitude'] = merged['longitude'].astype(float)
    merged.dropna(subset=['latitude', 'longitude', 'clean_zip'], inplace=True)
    return merged.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Optimal K recommendation
# ---------------------------------------------------------------------------
def recommend_k(zip_composites, chosen_k, min_cap, max_cap, hard_floor):
    if len(zip_composites) == 0 or zip_composites.sum() <= 0:
        return {"optimal_k": chosen_k, "k_min": 1, "k_max": chosen_k * 2}
    total = zip_composites.sum()
    nz = zip_composites[zip_composites > 0]
    max_zip = nz.max()
    k_max_from_cap = math.floor(max_cap * total / (max_zip * 1000)) if max_zip > 0 else chosen_k * 2
    k_max = min(k_max_from_cap, len(nz))
    k_min = max(1, math.ceil(chosen_k * 0.5))
    return {
        "optimal_k": int(chosen_k),
        "k_min": int(k_min),
        "k_max": int(k_max),
        "chosen_k_feasible": int(k_min) <= chosen_k <= int(k_max),
    }


# ---------------------------------------------------------------------------
# BIDIRECTIONAL centroid seeding
# ---------------------------------------------------------------------------
def seed_centroids_bidirectional(coords: np.ndarray, weights: np.ndarray, k: int) -> np.ndarray:
    """
    Seeds K territory centroids from both coasts converging inward.
    
    Why: US HCP distribution is coast-heavy. Seeding from both ends of the
    longitude spectrum ensures centre territories claim their fair share of
    sparse interior ZIPs before coastal territories expand inward.
    
    Algorithm:
    1. Sort nonzero ZIPs by longitude (west → east)
    2. Split into left half (western K/2) and right half (eastern K/2)
    3. Within each half, divide into equal-weight buckets by latitude
       (so north/south variation is also captured within each coast)
    4. Return weighted centroid of each bucket as starting centroid
    """
    nonzero_mask = weights > 0
    nz_coords = coords[nonzero_mask]
    nz_weights = weights[nonzero_mask]

    if len(nz_coords) < k:
        idx = np.random.choice(len(coords), size=k, replace=False)
        return coords[idx]

    # Sort by longitude west→east
    lon_order = np.argsort(nz_coords[:, 1])
    sorted_coords = nz_coords[lon_order]
    sorted_weights = nz_weights[lon_order]

    # Cumulative weight along longitude axis
    cum_w = np.cumsum(sorted_weights)
    total_w = cum_w[-1]

    # Split point: half the territories from left, half from right
    k_left = k // 2
    k_right = k - k_left

    # Find longitude split index (midpoint by weight)
    mid_w = total_w / 2.0
    split_idx = np.searchsorted(cum_w, mid_w)
    split_idx = max(k_left, min(split_idx, len(sorted_coords) - k_right))

    left_coords  = sorted_coords[:split_idx]
    left_weights = sorted_weights[:split_idx]
    right_coords  = sorted_coords[split_idx:]
    right_weights = sorted_weights[split_idx:]

    def bucket_centroids(bcoords, bweights, n_buckets):
        """Divide into n equal-weight buckets sorted by latitude, return centroids."""
        if len(bcoords) == 0 or n_buckets == 0:
            return []
        lat_order = np.argsort(bcoords[:, 0])
        bc = bcoords[lat_order]
        bw = bweights[lat_order]
        bw_cum = np.cumsum(bw)
        bw_total = bw_cum[-1]
        bucket_size = bw_total / n_buckets
        centroids = []
        prev = 0.0
        for i in range(n_buckets):
            target = bucket_size * (i + 1)
            if i == n_buckets - 1:
                mask = bw_cum > prev
            else:
                mask = (bw_cum > prev) & (bw_cum <= target)
            pts = bc[mask]
            ws  = bw[mask]
            if len(pts) == 0:
                # fallback: nearest available
                idx = min(int(prev / bw_total * len(bc)), len(bc) - 1)
                centroids.append(bc[idx])
            else:
                ws_sum = ws.sum()
                centroids.append(
                    np.average(pts, weights=ws, axis=0) if ws_sum > 0 else pts.mean(axis=0)
                )
            prev = target
        return centroids

    left_centroids  = bucket_centroids(left_coords,  left_weights,  k_left)
    right_centroids = bucket_centroids(right_coords, right_weights, k_right)

    all_centroids = left_centroids + right_centroids
    # Shuffle so KMeans doesn't bias on order
    random.seed(42)
    random.shuffle(all_centroids)
    return np.array(all_centroids)



# ---------------------------------------------------------------------------
# ZIP NEIGHBOR GRAPH  — used for contiguous assignment & rebalancing
# ---------------------------------------------------------------------------
def build_zip_neighbor_graph(coords: np.ndarray, n_neighbors: int = 6) -> dict:
    """
    For each ZIP index build a set of its n_neighbors nearest ZIP indices.
    Uses scipy cKDTree — O(n log n), ~5MB instead of O(n^2) distance matrix.
    Returns dict {zip_idx: set of neighbor zip_idxs}.
    """
    tree = cKDTree(coords)
    # Query n_neighbors+1 because the first result is always the point itself
    dists, indices = tree.query(coords, k=min(n_neighbors + 1, len(coords)))
    neighbors = {}
    for i, row in enumerate(indices):
        neighbors[i] = set(int(j) for j in row if j != i)
    return neighbors


# ---------------------------------------------------------------------------
# CONTIGUOUS ZIP ASSIGNMENT
# Uses a priority queue (min-heap) to grow territories outward from centroids,
# only assigning ZIPs that are geographically adjacent to already-claimed ZIPs.
# This prevents territories reaching across each other.
# ---------------------------------------------------------------------------
def assign_zips_contiguous(
    coords: np.ndarray,
    weights: np.ndarray,
    centroids: np.ndarray,
    min_cap: float,
    max_cap: float,
    hard_floor: float,
    neighbors: dict,
) -> tuple:
    k = len(centroids)
    n = len(coords)
    labels = np.full(n, -1, dtype=int)
    cluster_loads = np.zeros(k)
    locked = np.zeros(k, dtype=bool)  # territory locked once balanced

    nonzero_mask = weights > 0

    # Seed each territory with its single closest non-zero ZIP
    dist_to_centroids = cdist(coords, centroids, metric='euclidean')
    seeded = set()
    for t_id in range(k):
        order = np.argsort(dist_to_centroids[:, t_id])
        for z in order:
            if nonzero_mask[z] and z not in seeded:
                labels[z] = t_id
                cluster_loads[t_id] += weights[z]
                seeded.add(z)
                break

    # Priority queue: (distance_to_centroid, zip_idx, territory_id)
    # Each territory's frontier = unassigned neighbors of claimed ZIPs
    frontier = []  # min-heap by distance to centroid

    def add_frontier(z_idx, t_id):
        d = dist_to_centroids[z_idx, t_id]
        heapq.heappush(frontier, (d, z_idx, t_id))

    # Initialise frontier from seeds
    for z_idx in range(n):
        if labels[z_idx] >= 0:
            for nb in neighbors[z_idx]:
                if labels[nb] == -1:
                    add_frontier(nb, labels[z_idx])

    # Grow territories greedily (Prim-like expansion)
    while frontier:
        dist, z_idx, t_id = heapq.heappop(frontier)
        if labels[z_idx] != -1:
            continue  # already assigned
        if locked[t_id]:
            continue  # territory full, skip

        w = weights[z_idx]
        # Accept if territory not over cap
        if cluster_loads[t_id] + w <= max_cap:
            labels[z_idx] = t_id
            cluster_loads[t_id] += w
            # Lock territory once it hits min_cap
            if cluster_loads[t_id] >= min_cap:
                locked[t_id] = True
            # Expand frontier
            for nb in neighbors[z_idx]:
                if labels[nb] == -1:
                    add_frontier(nb, t_id)
        else:
            # Try other territories for this ZIP via normal fallback below
            pass

    # Any remaining unassigned ZIPs — assign to nearest territory that isn't locked
    unassigned = np.where(labels == -1)[0]
    for z_idx in unassigned:
        order = np.argsort(dist_to_centroids[z_idx])
        for t_id in order:
            if cluster_loads[t_id] + weights[z_idx] <= max_cap:
                labels[z_idx] = t_id
                cluster_loads[t_id] += weights[z_idx]
                break
        else:
            # Force-assign to closest
            t_id = int(np.argmin(dist_to_centroids[z_idx]))
            labels[z_idx] = t_id
            cluster_loads[t_id] += weights[z_idx]

    return labels, cluster_loads


# ---------------------------------------------------------------------------
# CONTIGUOUS REBALANCE
# Your idea: work from extremes outward, donate border ZIPs to neighbors,
# lock territories once balanced. Only moves ZIPs along shared borders.
# ---------------------------------------------------------------------------
def contiguous_rebalance(
    coords: np.ndarray,
    weights: np.ndarray,
    labels: np.ndarray,
    cluster_loads: np.ndarray,
    k: int,
    hard_floor: float,
    min_cap: float,
    max_cap: float,
    neighbors: dict,
    max_rounds: int = 25,
) -> tuple:
    """
    Rebalance by donation along geographic borders only.

    Algorithm per round:
      1. Sort territories by load (most over-cap first, then most under-floor last)
      2. For each over-cap territory O:
         a. Find its border ZIPs (ZIPs that have a neighbor in a different territory)
         b. Offer border ZIPs to adjacent under/low territories (closest centroid first)
         c. Lock O once it drops into range; lock receiving territory once in range
      3. For each under-floor territory U:
         a. Find adjacent territories with spare capacity
         b. Accept their border ZIPs closest to U's centroid
         c. Lock U once it clears the floor
      4. Repeat until no locked changes occur or max_rounds hit
    """
    labels = labels.copy()
    cluster_loads = cluster_loads.copy()
    locked = np.zeros(k, dtype=bool)
    nonzero_mask = weights > 0

    def centroid(t_id):
        idx = np.where((labels == t_id) & nonzero_mask)[0]
        if len(idx) == 0:
            return coords.mean(axis=0)
        return coords[idx].mean(axis=0)

    def border_zips(t_id):
        """ZIPs in t_id that have at least one neighbor in a different territory."""
        t_idx = np.where((labels == t_id) & nonzero_mask)[0]
        border = []
        for z in t_idx:
            for nb in neighbors[z]:
                if labels[nb] != t_id:
                    border.append(z)
                    break
        return border

    for round_num in range(max_rounds):
        loads = np.array([cluster_loads[t] for t in range(k)])
        made_progress = False

        # ── Step 1: Bleed over-cap territories into neighbours ──
        overcap = sorted(
            [t for t in range(k) if loads[t] > max_cap and not locked[t]],
            key=lambda t: -loads[t]  # most over first
        )

        for o_id in overcap:
            if locked[o_id]:
                continue
            bz = border_zips(o_id)
            if not bz:
                continue

            # Sort by distance to centroid of most-needy adjacent territory
            # Find adjacent territories
            adj_ids = set()
            for z in bz:
                for nb in neighbors[z]:
                    if labels[nb] != o_id:
                        adj_ids.add(labels[nb])

            # Prefer to donate to under-floor first, then under-min_cap
            adj_sorted = sorted(
                adj_ids,
                key=lambda t: (
                    0 if loads[t] < hard_floor else
                    1 if loads[t] < min_cap else
                    2 if loads[t] <= max_cap else 3
                )
            )

            o_center = centroid(o_id)
            for recv_id in adj_sorted:
                if loads[o_id] <= max_cap:
                    break
                if locked[recv_id]:
                    continue
                recv_load = loads[recv_id]
                if recv_load >= max_cap:
                    continue  # receiver also full

                # ZIPs closest to receiver centroid
                recv_center = centroid(recv_id)
                candidate_zips = [z for z in bz if labels[z] == o_id]
                candidate_zips.sort(
                    key=lambda z: np.linalg.norm(coords[z] - recv_center)
                )

                for z in candidate_zips:
                    w = weights[z]
                    if w == 0:
                        continue
                    new_o_load = loads[o_id] - w
                    new_r_load = recv_load + w
                    if new_o_load < hard_floor:
                        continue  # donor would drop below floor
                    if new_r_load > max_cap:
                        continue  # receiver would overflow

                    labels[z] = recv_id
                    loads[o_id] = new_o_load
                    loads[recv_id] = new_r_load
                    cluster_loads[o_id] = new_o_load
                    cluster_loads[recv_id] = new_r_load
                    recv_load = new_r_load
                    bz = [x for x in bz if x != z]
                    made_progress = True

                    if hard_floor <= loads[recv_id] <= max_cap:
                        locked[recv_id] = True
                    if hard_floor <= loads[o_id] <= max_cap:
                        locked[o_id] = True
                        break

        # ── Step 2: Pull ZIPs into under-floor territories ──
        underfloor = sorted(
            [t for t in range(k) if loads[t] < hard_floor and not locked[t]],
            key=lambda t: loads[t]  # most starved first
        )

        for u_id in underfloor:
            if locked[u_id]:
                continue
            u_center = centroid(u_id)
            u_load = loads[u_id]

            # Find adjacent territories
            u_border_neighbors = set()
            u_idx = np.where((labels == u_id) & nonzero_mask)[0]
            for z in u_idx:
                for nb in neighbors[z]:
                    if labels[nb] != u_id:
                        u_border_neighbors.add(labels[nb])

            # Sort donors by load descending (most over first)
            donors = sorted(
                u_border_neighbors,
                key=lambda t: -loads[t]
            )

            for donor_id in donors:
                if u_load >= hard_floor:
                    break
                if locked[donor_id]:
                    continue
                if loads[donor_id] <= hard_floor:
                    continue  # donor can't afford to give

                # Donor's border ZIPs closest to u_center
                donor_border = border_zips(donor_id)
                donor_border.sort(
                    key=lambda z: np.linalg.norm(coords[z] - u_center)
                )

                for z in donor_border:
                    w = weights[z]
                    if w == 0:
                        continue
                    if loads[donor_id] - w < hard_floor:
                        continue
                    if u_load + w > max_cap:
                        continue

                    labels[z] = u_id
                    loads[u_id] = u_load + w
                    loads[donor_id] -= w
                    cluster_loads[u_id] = loads[u_id]
                    cluster_loads[donor_id] = loads[donor_id]
                    u_load += w
                    made_progress = True

                    if loads[u_id] >= hard_floor:
                        locked[u_id] = True
                        break

        if not made_progress:
            break

    return labels, cluster_loads

# ---------------------------------------------------------------------------
# Constrained ZIP assignment
# ---------------------------------------------------------------------------
def assign_zips(coords, weights, centroids, min_cap, max_cap, hard_floor=650.0):
    k = len(centroids)
    dist_matrix = cdist(coords, centroids, metric='euclidean')
    labels = np.full(len(coords), -1, dtype=int)
    cluster_loads = np.zeros(k)

    nonzero_idx = np.where(weights > 0)[0]
    zero_idx    = np.where(weights == 0)[0]

    # Sort nonzero ZIPs by distance to their nearest centroid (greedy nearest-first)
    closest = np.argmin(dist_matrix[nonzero_idx], axis=1)
    dist_to_closest = dist_matrix[nonzero_idx, closest]
    order = np.argsort(dist_to_closest)

    for pos in order:
        zip_idx = nonzero_idx[pos]
        w = weights[zip_idx]
        territory_order = np.argsort(dist_matrix[zip_idx])
        for t_id in territory_order:
            if cluster_loads[t_id] + w <= max_cap:
                labels[zip_idx] = t_id
                cluster_loads[t_id] += w
                break
        else:
            # No room anywhere — assign to closest regardless
            t_id = territory_order[0]
            labels[zip_idx] = t_id
            cluster_loads[t_id] += w

    for zip_idx in zero_idx:
        labels[zip_idx] = np.argmin(dist_matrix[zip_idx])

    return labels, cluster_loads


# ---------------------------------------------------------------------------
# BORDER DONATION rebalancer (replaces dissolve-and-rebalance)
# ---------------------------------------------------------------------------
def border_donation_rebalance(
    coords: np.ndarray,
    weights: np.ndarray,
    labels: np.ndarray,
    cluster_loads: np.ndarray,
    k: int,
    hard_floor: float,
    max_cap: float,
    max_rounds: int = 15,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    For each underweight territory, ask neighbouring territories to donate
    their border ZIPs one by one until the underweight territory reaches
    hard_floor — without the donor dropping below hard_floor itself.

    No territories are dissolved. K is always exactly maintained.
    No ghost territories possible (each territory always has nonzero ZIPs
    after donation, since we only take from donors that stay above floor).

    Process per round:
      1. Find all territories below hard_floor (underweight)
      2. For each underweight territory U:
         a. Find all ZIPs belonging to neighbours that are geographically
            closest to U's centroid
         b. Donate one ZIP at a time from the most-overweight neighbour
            whose donation keeps donor >= hard_floor
         c. Stop when U >= hard_floor or no more donatable ZIPs
      3. Repeat until no underweight territories remain or max_rounds hit
    """
    labels = labels.copy()
    cluster_loads = cluster_loads.copy()

    nonzero_mask = weights > 0

    for round_num in range(max_rounds):
        # Find underweight territories (only counting nonzero-weight ZIPs)
        underweight = []
        for t_id in range(k):
            nz_in_t = np.where((labels == t_id) & nonzero_mask)[0]
            load = weights[nz_in_t].sum()
            if load < hard_floor and len(nz_in_t) > 0:
                underweight.append(t_id)

        if not underweight:
            break

        made_progress = False

        for u_id in underweight:
            u_nz = np.where((labels == u_id) & nonzero_mask)[0]
            if len(u_nz) == 0:
                continue
            u_center = coords[u_nz].mean(axis=0)
            u_load = weights[u_nz].sum()

            # Find neighbouring territories: those that have a nonzero ZIP
            # within a reasonable distance of U's centroid
            # "neighbour" = any territory whose centroid is within top-6 nearest
            neighbour_ids = set()
            for t_id in range(k):
                if t_id == u_id:
                    continue
                t_nz = np.where((labels == t_id) & nonzero_mask)[0]
                if len(t_nz) == 0:
                    continue
                neighbour_ids.add(t_id)

            # Sort neighbours by centroid distance to U
            neighbour_dists = []
            for n_id in neighbour_ids:
                n_nz = np.where((labels == n_id) & nonzero_mask)[0]
                n_center = coords[n_nz].mean(axis=0)
                dist = np.linalg.norm(n_center - u_center)
                neighbour_dists.append((n_id, dist))
            neighbour_dists.sort(key=lambda x: x[1])
            # Consider closest 8 neighbours
            candidate_neighbours = [n for n, _ in neighbour_dists[:8]]

            for n_id in candidate_neighbours:
                if u_load >= hard_floor:
                    break

                n_nz = np.where((labels == n_id) & nonzero_mask)[0]
                n_load = weights[n_nz].sum()

                # Donor must stay above hard_floor after donation
                # Find border ZIPs of donor: nonzero ZIPs in donor sorted by 
                # distance to U's centroid (closest = border)
                dists_to_u = np.linalg.norm(coords[n_nz] - u_center, axis=1)
                donation_order = np.argsort(dists_to_u)  # closest to U first

                for pos in donation_order:
                    zip_idx = n_nz[pos]
                    w = weights[zip_idx]

                    # Check donor stays above floor after giving this ZIP
                    if n_load - w < hard_floor:
                        continue  # donor can't afford this donation
                    # Check receiver won't exceed max_cap
                    if u_load + w > max_cap:
                        continue  # would make U too heavy

                    # Donate
                    labels[zip_idx] = u_id
                    cluster_loads[u_id] += w
                    cluster_loads[n_id] -= w
                    u_load += w
                    n_load -= w
                    made_progress = True

                    if u_load >= hard_floor:
                        break

        if not made_progress:
            break

    return labels, cluster_loads


# ---------------------------------------------------------------------------
# Statistics per territory
# ---------------------------------------------------------------------------
def compute_stats(df, labels, weights, k, min_cap, max_cap, hard_floor):
    stats = {}
    for t_id in range(k):
        t_mask = labels == t_id
        total_w = int(weights[t_mask].sum())
        zip_count = int(t_mask.sum())
        nz_count = int(((labels == t_id) & (weights > 0)).sum())

        t_df = df[t_mask]
        diameter = 0
        if len(t_df) >= 2:
            try:
                mins = t_df[['latitude', 'longitude']].min()
                maxs = t_df[['latitude', 'longitude']].max()
                diameter = int(geodesic(
                    (mins.latitude, mins.longitude),
                    (maxs.latitude, maxs.longitude)
                ).miles)
            except Exception:
                pass

        if total_w < hard_floor:
            status, msg = "Red", f"Below floor — needs {int(hard_floor - total_w)} more"
        elif total_w < min_cap:
            status, msg = "Yellow", f"Needs {int(min_cap - total_w)} more"
        elif total_w > max_cap:
            status, msg = "Orange", f"Over cap by {int(total_w - max_cap)}"
        else:
            status, msg = "Green", "Optimal"

        stats[t_id] = {
            "Status": status,
            "Message": msg,
            "Weight": total_w,
            "ZipCount": zip_count,
            "NonZeroZips": nz_count,
            "Diameter": diameter,
        }
    return stats


# ---------------------------------------------------------------------------
# CONVEX HULL territory shapes for map
# ---------------------------------------------------------------------------
def territory_shapes(df: pd.DataFrame, labels: np.ndarray,
                      weights: np.ndarray, k: int) -> dict:
    """
    Returns a dict {t_id: list of rings [[lat,lon],...]} using Voronoi-based
    territory boundaries — each ZIP gets a Voronoi cell clipped to data extent,
    then cells are merged by territory. Falls back to convex hull if shapely unavailable.
    """
    from scipy.spatial import Voronoi, ConvexHull

    all_coords = df[['latitude', 'longitude']].values

    # Try Voronoi approach first
    if HAS_SHAPELY and len(all_coords) >= 4:
        try:
            lat_min, lat_max = all_coords[:,0].min()-1, all_coords[:,0].max()+1
            lon_min, lon_max = all_coords[:,1].min()-1, all_coords[:,1].max()+1
            bounds = shapely_box(lon_min, lat_min, lon_max, lat_max)

            # Mirror points to close edge cells
            margin = 3.0; n = 25
            mirrors = np.vstack([
                np.column_stack([np.linspace(lon_min-margin, lon_max+margin, n), np.full(n, lat_max+margin)]),
                np.column_stack([np.linspace(lon_min-margin, lon_max+margin, n), np.full(n, lat_min-margin)]),
                np.column_stack([np.full(n, lon_min-margin), np.linspace(lat_min-margin, lat_max+margin, n)]),
                np.column_stack([np.full(n, lon_max+margin), np.linspace(lat_min-margin, lat_max+margin, n)]),
            ])
            all_pts = np.vstack([all_coords[:, ::-1], mirrors])  # lon,lat
            vor = Voronoi(all_pts)

            zip_polys = {}
            for i in range(len(all_coords)):
                reg = vor.regions[vor.point_region[i]]
                if -1 in reg or len(reg) == 0:
                    continue
                try:
                    poly = ShapelyPolygon(vor.vertices[reg]).intersection(bounds)
                    if not poly.is_empty:
                        zip_polys[i] = poly
                except Exception:
                    continue

            shapes = {}
            for t_id in range(k):
                t_idx = np.where(labels == t_id)[0]
                polys = [zip_polys[i] for i in t_idx if i in zip_polys]
                if not polys:
                    continue
                try:
                    merged = unary_union(polys)
                    if merged.is_empty:
                        continue
                    if merged.geom_type == 'Polygon':
                        rings = [[[lat, lon] for lon, lat in merged.exterior.coords]]
                    else:
                        rings = [[[lat, lon] for lon, lat in p.exterior.coords] for p in merged.geoms]
                    shapes[t_id] = rings[0] if len(rings) == 1 else rings[0]
                    # Store all rings
                    if len(rings) > 1:
                        shapes[t_id] = rings
                except Exception:
                    continue
            if shapes:
                return shapes
        except Exception:
            pass

    # Fallback: convex hull per territory
    shapes = {}
    for t_id in range(k):
        nz_idx = np.where((labels == t_id) & (weights > 0))[0]
        if len(nz_idx) < 3:
            continue
        pts = df.iloc[nz_idx][['latitude', 'longitude']].values
        try:
            hull = ConvexHull(pts)
            ring = pts[hull.vertices].tolist()
            ring.append(ring[0])
            shapes[t_id] = ring
        except Exception:
            pass
    return shapes


# Aliases
territory_alpha_shapes = territory_shapes
territory_convex_hulls = territory_shapes


# ---------------------------------------------------------------------------
# TERRITORY COLORS — visually distinct palette
# ---------------------------------------------------------------------------
def assign_region(lat: float, lon: float) -> int:
    """Assign a ZIP centroid to one of 4 geographic regions."""
    # West: lon < -104
    if lon < -104:
        return 0
    # Central: -104 <= lon < -88
    elif lon < -88:
        return 1
    # Southeast: lon >= -88 and lat < 37
    elif lat < 37:
        return 3
    # Northeast: lon >= -88 and lat >= 37
    else:
        return 2


# 5 distinct shades per region (West, Central, Northeast, Southeast)
REGION_PALETTES = [
    # West — blues
    ["#1a4fa0", "#2d7dd2", "#56aef7", "#a8d4f5", "#d0eaff"],
    # Central — greens
    ["#1a6b3a", "#2e9e55", "#52c278", "#96dba8", "#c8f0d4"],
    # Northeast — purples
    ["#4a1a8a", "#7b2fbf", "#a855f7", "#c89af5", "#e8d5ff"],
    # Southeast — oranges
    ["#8a3a00", "#c45c00", "#f08030", "#f5b080", "#fad8b8"],
]


def generate_territory_colors(k: int) -> list:
    """Return k colors — placeholder, replaced by region-aware coloring in map gen."""
    # Fallback for non-map uses
    import colorsys
    colors = []
    for i in range(k):
        h = (i * 137.5) % 360 / 360
        r, g, b = colorsys.hls_to_rgb(h, 0.52, 0.65)
        colors.append('#{:02x}{:02x}{:02x}'.format(int(r*255), int(g*255), int(b*255)))
    return colors


# ---------------------------------------------------------------------------
# Map generation — dots only (outline ZIPs define territory boundary)
# ---------------------------------------------------------------------------
def generate_map_html(df: pd.DataFrame, stats: dict, k: int, title: str) -> str:
    center_lat = df['latitude'].mean() if not df.empty else 39.8
    center_lon = df['longitude'].mean() if not df.empty else -98.5

    m = folium.Map(location=[center_lat, center_lon], zoom_start=5,
                   prefer_canvas=True, tiles='CartoDB positron')
    m.get_root().html.add_child(
        folium.Element(
            f'<h3 align="center" style="font-size:14px;font-family:sans-serif;'
            f'margin:6px 0;color:#333"><b>{title}</b></h3>'
        )
    )

    labels_arr  = df['Territory_ID'].values
    weights_arr = df['final_weight'].values
    coords      = df[['latitude','longitude']].values
    status_colors = {"Green": "#1aa15a", "Yellow": "#d18a17",
                     "Red": "#d13b3b", "Orange": "#e07020"}

    active_df = df[df['final_weight'] > 0].copy()

    # ── Assign region-aware colors (5 shades × 4 regions) ────────────────
    # Find weighted centroid of each territory → assign region → pick shade
    territory_colors = {}
    region_shade_count = [0, 0, 0, 0]  # shade index per region
    for t_id in range(k):
        t_idx = np.where(labels_arr == t_id)[0]
        if len(t_idx) == 0:
            territory_colors[t_id] = "#888888"
            continue
        w = weights_arr[t_idx]
        c = coords[t_idx]
        wc = (c * w[:,None]).sum(0) / w.sum() if w.sum() > 0 else c.mean(0)
        region = assign_region(float(wc[0]), float(wc[1]))
        shade_idx = region_shade_count[region] % len(REGION_PALETTES[region])
        territory_colors[t_id] = REGION_PALETTES[region][shade_idx]
        region_shade_count[region] += 1

    # ── Identify border ZIPs (have a neighbor in a different territory) ──
    from scipy.spatial import cKDTree
    tree = cKDTree(coords)
    _, nn_idx = tree.query(coords, k=min(8, len(coords)))
    border_set = set()
    for i in range(len(coords)):
        for j in nn_idx[i]:
            if j != i and labels_arr[j] != labels_arr[i]:
                border_set.add(i)
                break

    # ── Convex hull outline (no fill) for territory boundary ────────────────
    from scipy.spatial import ConvexHull
    outline_group = folium.FeatureGroup(name="Territory boundaries", show=True)
    for t_id in range(k):
        t_idx = np.where((labels_arr == t_id) & (weights_arr > 0))[0]
        if len(t_idx) < 3:
            continue
        color = territory_colors.get(t_id, "#888")
        pts = coords[t_idx]
        try:
            hull = ConvexHull(pts)
            hull_pts = pts[hull.vertices].tolist()
            hull_pts.append(hull_pts[0])
            folium.Polygon(
                locations=hull_pts,
                color=color,
                weight=2.0,
                fill=False,          # outline only — no fill
                opacity=0.85,
                tooltip=f"T{t_id}",
            ).add_to(outline_group)
        except Exception:
            pass
    outline_group.add_to(m)

    # ── ZIP dot markers — border ZIPs larger ─────────────────────────────
    dot_group = folium.FeatureGroup(name="ZIP markers", show=True)
    for df_idx, row in active_df.iterrows():
        t_id = int(row['Territory_ID'])
        color = territory_colors.get(t_id, "#888")
        # Use original df positional index to check border_set
        pos_idx = df.index.get_loc(df_idx) if hasattr(df.index, 'get_loc') else df_idx
        is_border = pos_idx in border_set
        folium.CircleMarker(
            location=[row['latitude'], row['longitude']],
            radius=5 if is_border else 3,
            color=color,
            fill=True,
            fill_color=color,
            fill_opacity=1.0 if is_border else 0.75,
            weight=0.5 if is_border else 0,
            tooltip=f"T{t_id} | ZIP: {row['clean_zip']} | Index: {int(row['final_weight'])}",
        ).add_to(dot_group)
    dot_group.add_to(m)

    # ── Territory label markers ──
    label_group = folium.FeatureGroup(name="Territory labels", show=True)
    for t_id in range(k):
        t_data = active_df[active_df['Territory_ID'] == t_id]
        if t_data.empty:
            continue
        s = stats.get(t_id, {})
        status = s.get("Status", "")
        bg = status_colors.get(status, "#555")
        center = [t_data['latitude'].mean(), t_data['longitude'].mean()]
        folium.Marker(
            center,
            icon=folium.DivIcon(
                html=(
                    f'<div style="font-size:7.5pt;font-weight:700;color:#fff;'
                    f'background:{bg};padding:2px 5px;border-radius:4px;'
                    f'white-space:nowrap;box-shadow:0 1px 3px rgba(0,0,0,.3)">'
                    f'T{t_id}&nbsp;|&nbsp;{s.get("Weight")}</div>'
                ),
                icon_size=(88, 20),
                icon_anchor=(44, 10),
            ),
            popup=folium.Popup(
                f"<b>Territory {t_id}</b><br>"
                f"Status: <b>{s.get('Status')}</b><br>"
                f"Index: {s.get('Weight')}<br>"
                f"Active ZIPs: {s.get('NonZeroZips')}<br>"
                f"Diameter: {s.get('Diameter')} mi<br>"
                f"Centroid: {s.get('Centroid_Lat','?')}, {s.get('Centroid_Lon','?')}<br>"
                f"{s.get('Message')}",
                max_width=220,
            ),
        ).add_to(label_group)
    label_group.add_to(m)

    folium.LayerControl(collapsed=False).add_to(m)
    return m.get_root().render()


# ---------------------------------------------------------------------------
# Master run_scenario
# ---------------------------------------------------------------------------
def _bisect_for_groups(coords, weights, indices, k):
    """Recursive geographic bisection — splits by longest axis, balanced by weight."""
    if k == 1:
        return [indices]
    sc   = coords[indices]
    axis = 0 if (sc[:,0].max()-sc[:,0].min()) >= (sc[:,1].max()-sc[:,1].min()) else 1
    so   = np.argsort(sc[:, axis])
    si   = indices[so]; sw = weights[si]
    cu   = np.cumsum(sw)
    cut  = max(1, min(int(np.searchsorted(cu, cu[-1] / 2.0)), len(si)-1))
    kl   = k // 2; kr = k - kl
    return (_bisect_for_groups(coords, weights, si[:cut], kl) +
            _bisect_for_groups(coords, weights, si[cut:], kr))


def run_scenario(df, k, min_cap, max_cap, hard_floor=650.0):
    """
    Territory generation pipeline:
      1. Normalise weights to sum = k * 1000
      2. Recursive geographic bisection → initial group assignments
      3. Three-pass border-donation rebalancing:
         Pass 1 — drain over-cap (relaxed donor floor)
         Pass 2 — lift under-floor
         Pass 3 — tighten yellow (below min_cap)
      4. Compute stats + weighted centroids
    """
    scenario_df = df.copy().reset_index(drop=True)

    total_composite = scenario_df['zip_composite'].sum()
    if total_composite <= 0:
        raise ValueError("Total composite score is zero — check column names match file.")

    scenario_df['final_weight'] = (
        (scenario_df['zip_composite'] / total_composite) * k * 1000
    ).fillna(0).clip(lower=0)

    coords  = scenario_df[['latitude', 'longitude']].values
    weights = scenario_df['final_weight'].values.astype(float)
    scenario_df['final_weight'] = weights

    nonzero_mask = weights > 0
    if nonzero_mask.sum() < k:
        raise ValueError(
            f"Not enough non-zero ZIPs ({nonzero_mask.sum()}) to form {k} territories."
        )

    zip_composites = scenario_df['zip_composite'].values
    k_rec = recommend_k(zip_composites, k, min_cap, max_cap, hard_floor)

    # ── Step 1: Recursive bisection ──────────────────────────────────────
    print(f"Bisecting {len(coords)} ZIPs into {k} groups...")
    groups = _bisect_for_groups(coords, weights, np.arange(len(scenario_df)), k)
    labels = np.zeros(len(scenario_df), dtype=int)
    for t_id, grp in enumerate(groups):
        labels[grp] = t_id
    gc.collect()

    # ── Step 2: Build ZIP neighbor graph ─────────────────────────────────
    print("Building neighbor graph...")
    zip_tree = cKDTree(coords)
    _, nn    = zip_tree.query(coords, k=min(8, len(coords)))
    neighbors = {i: set(int(j) for j in nn[i] if j != i) for i in range(len(coords))}
    gc.collect()

    # Helpers
    def get_loads():
        return np.array([weights[labels == t].sum() for t in range(k)])

    def border_zips_of(t_id):
        return [z for z in np.where(labels == t_id)[0]
                if any(labels[nb] != t_id for nb in neighbors[z])]

    def wcentroid(t_id):
        idx = np.where(labels == t_id)[0]
        w   = weights[idx]
        return (coords[idx] * w[:,None]).sum(0) / w.sum() if len(idx) > 0 else coords.mean(0)

    # ── Step 3a: Drain over-cap (relaxed donor floor = hard_floor * 0.5) ─
    print("Pass 1: draining over-cap territories...")
    relaxed_floor = hard_floor * 0.5
    for rnd in range(200):
        loads = get_loads()
        over  = sorted([t for t in range(k) if loads[t] > max_cap], key=lambda t: -loads[t])
        if not over:
            print(f"  Over-cap cleared at round {rnd}")
            break
        made = False
        for o_id in over:
            if loads[o_id] <= max_cap:
                continue
            bz  = sorted(border_zips_of(o_id), key=lambda z: weights[z])
            adj = set(labels[nb] for z in bz for nb in neighbors[z] if labels[nb] != o_id)
            for recv_id in sorted(adj, key=lambda t: loads[t]):
                if loads[o_id] <= max_cap:
                    break
                if loads[recv_id] >= max_cap:
                    continue
                rc = wcentroid(recv_id)
                for z in sorted(bz, key=lambda z: np.linalg.norm(coords[z] - rc)):
                    w = weights[z]
                    if loads[recv_id] + w > max_cap:
                        continue
                    if loads[o_id] - w < relaxed_floor:
                        continue
                    labels[z]      = recv_id
                    loads[o_id]   -= w
                    loads[recv_id] += w
                    bz = [x for x in bz if x != z]
                    made = True
                    break
        if not made:
            print(f"  Pass 1 stalled at round {rnd}")
            break

    # ── Step 3b: Lift under-floor territories ─────────────────────────────
    print("Pass 2: lifting under-floor territories...")
    for rnd in range(200):
        loads = get_loads()
        under = sorted([t for t in range(k) if loads[t] < hard_floor], key=lambda t: loads[t])
        if not under:
            print(f"  Under-floor cleared at round {rnd}")
            break
        made = False
        for u_id in under:
            if loads[u_id] >= hard_floor:
                continue
            uc  = wcentroid(u_id)
            adj = set(labels[nb] for z in np.where(labels == u_id)[0]
                      for nb in neighbors[z] if labels[nb] != u_id)
            for d_id in sorted(adj, key=lambda t: -loads[t]):
                if loads[u_id] >= hard_floor:
                    break
                if loads[d_id] <= hard_floor + 50:
                    continue
                for z in sorted(border_zips_of(d_id), key=lambda z: np.linalg.norm(coords[z] - uc)):
                    w = weights[z]
                    if loads[d_id] - w < hard_floor:
                        continue
                    if loads[u_id] + w > max_cap:
                        continue
                    labels[z]      = u_id
                    loads[u_id]   += w
                    loads[d_id]   -= w
                    made = True
                    break
        if not made:
            print(f"  Pass 2 stalled at round {rnd}")
            break

    # ── Step 3c: Tighten yellow (below min_cap) ───────────────────────────
    print("Pass 3: tightening yellow territories...")
    for rnd in range(100):
        loads = get_loads()
        yellow = sorted([t for t in range(k) if loads[t] < min_cap], key=lambda t: loads[t])
        if not yellow:
            print(f"  Yellow cleared at round {rnd}")
            break
        made = False
        for u_id in yellow:
            if loads[u_id] >= min_cap:
                continue
            uc  = wcentroid(u_id)
            adj = set(labels[nb] for z in np.where(labels == u_id)[0]
                      for nb in neighbors[z] if labels[nb] != u_id)
            for d_id in sorted(adj, key=lambda t: -loads[t]):
                if loads[u_id] >= min_cap:
                    break
                if loads[d_id] <= min_cap + 20:
                    continue
                for z in sorted(border_zips_of(d_id), key=lambda z: np.linalg.norm(coords[z] - uc)):
                    w = weights[z]
                    if loads[d_id] - w < hard_floor:
                        continue
                    if loads[u_id] + w > max_cap:
                        continue
                    labels[z]      = u_id
                    loads[u_id]   += w
                    loads[d_id]   -= w
                    made = True
                    break
        if not made:
            print(f"  Pass 3 stalled at round {rnd}")
            break

    # ── Step 4: Finalise ──────────────────────────────────────────────────
    scenario_df['Territory_ID'] = labels
    loads = get_loads()

    # Weighted centroids
    centroids_latlon = {}
    for t_id in range(k):
        idx = np.where(labels == t_id)[0]
        if len(idx) > 0:
            w = weights[idx]
            wc = (coords[idx] * w[:,None]).sum(0) / w.sum()
            centroids_latlon[t_id] = {
                "lat": round(float(wc[0]), 5),
                "lon": round(float(wc[1]), 5),
            }

    stats = compute_stats(scenario_df, labels, weights, k, min_cap, max_cap, hard_floor)
    for t_id, c in centroids_latlon.items():
        if t_id in stats:
            stats[t_id]["Centroid_Lat"] = c["lat"]
            stats[t_id]["Centroid_Lon"] = c["lon"]

    g = sum(1 for s in stats.values() if s["Status"] == "Green")
    y = sum(1 for s in stats.values() if s["Status"] == "Yellow")
    r = sum(1 for s in stats.values() if s["Status"] == "Red")
    o = sum(1 for s in stats.values() if s["Status"] == "Orange")
    print(f"Final: Green={g}, Yellow={y}, Red={r}, Over={o} | "
          f"Range {loads.min():.0f}–{loads.max():.0f}")

    return scenario_df, stats, k_rec


# ---------------------------------------------------------------------------
# AI inference via Groq API
# ---------------------------------------------------------------------------
def generate_ai_inference(stats: dict, k: int, tolerance_pct: int,
                           hard_floor: int, min_cap: int, max_cap: int) -> dict:
    """
    Call Claude to generate:
    - Executive narrative (3-4 sentences)
    - Per-outlier suggestions (Red/Orange/Yellow territories)
    - One strategic recommendation
    """
    try:
        client = Groq()

        # Summarise stats for the prompt
        green  = [t for t, s in stats.items() if s['Status'] == 'Green']
        yellow = [t for t, s in stats.items() if s['Status'] == 'Yellow']
        orange = [t for t, s in stats.items() if s['Status'] == 'Orange']
        red    = [t for t, s in stats.items() if s['Status'] == 'Red']

        outlier_detail = []
        for t_id in sorted(red + orange + yellow):
            s = stats[t_id]
            outlier_detail.append(
                f"T{t_id}: status={s['Status']}, index={s['Weight']}, "
                f"active_zips={s['NonZeroZips']}, diameter={s['Diameter']}mi, note={s['Message']}"
            )

        pct_green = round(len(green)/k*100) if k > 0 else 0
        prompt = f"""You are a senior pharmaceutical sales operations consultant writing a territory design review for a VP of Sales.

Territory alignment results:
- Total territories: {k}
- Index target: 1,000 per territory (±{tolerance_pct}% tolerance = {min_cap}–{max_cap} acceptable range)
- Hard floor (minimum viable): {hard_floor}
- GREEN — optimal ({min_cap}–{max_cap}): {len(green)} territories ({pct_green}% of total)
- YELLOW — slightly under ({hard_floor}–{min_cap}): {len(yellow)} territories
- ORANGE — over capacity (>{max_cap}): {len(orange)} territories
- RED — below hard floor (<{hard_floor}): {len(red)} territories

Outlier detail:
{chr(10).join(outlier_detail) if outlier_detail else "None — all territories are within the acceptable range."}

Write a professional territory design assessment. Be specific with numbers. Do NOT use generic filler language.

Respond ONLY with valid JSON, no markdown, no preamble:
{{
  "executive_summary": "Write 4-5 sentences. Lead with the overall balance score ({pct_green}% green). Name specific outlier territories by ID. Explain what is driving the imbalance (geographic concentration, data sparsity, or metric skew). State whether this alignment is deployment-ready or needs revision.",
  "outlier_suggestions": [
    {{"territory": "T8", "issue": "Specific diagnosis — e.g. over-concentrated Rx volume in 3 dense metro ZIPs pulling index to 1,340", "suggestion": "Specific fix — e.g. split T8 along the I-95 corridor, reassigning the Miami Beach ZIPs to adjacent T12 which is currently at 820"}}
  ],
  "strategic_recommendation": "Write 3-4 sentences. Give a concrete recommendation: whether to increase K, tighten tolerance, adjust segment call frequencies, or accept the current design. Reference the specific numbers. End with a deployment recommendation."
}}"""

        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        text = response.choices[0].message.content.strip()
        # Strip markdown fences if present
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        return json.loads(text.strip())
    except Exception as e:
        print(f"AI inference error: {e}")
        return {
            "executive_summary": f"Territory optimization completed with K={k}. "
                                  f"Green: {sum(1 for s in stats.values() if s['Status']=='Green')}, "
                                  f"Yellow: {sum(1 for s in stats.values() if s['Status']=='Yellow')}, "
                                  f"Orange: {sum(1 for s in stats.values() if s['Status']=='Orange')}, "
                                  f"Red: {sum(1 for s in stats.values() if s['Status']=='Red')}.",
            "outlier_suggestions": [],
            "strategic_recommendation": "Review outlier territories and consider adjusting tolerance or K."
        }


# ---------------------------------------------------------------------------
# EXCEL generation — rich formatted workbook
# ---------------------------------------------------------------------------
def _cell_fill(color_hex: str) -> PatternFill:
    return PatternFill("solid", fgColor=color_hex.lstrip('#'))

def _thin_border() -> Border:
    s = Side(style='thin', color='D0D7E0')
    return Border(left=s, right=s, top=s, bottom=s)

STATUS_FILLS = {
    "Green":  ("E3F5EC", "0E6E3C"),
    "Yellow": ("FDF3DF", "7A5000"),
    "Orange": ("FFF0E0", "7A3800"),
    "Red":    ("FBE4E4", "99231F"),
}

def build_excel(df_res: pd.DataFrame, stats: dict, k_rec: dict,
                k: int, tolerance_pct: int, hard_floor: int,
                min_cap: int, max_cap: int, ai: dict) -> bytes:
    wb = openpyxl.Workbook()

    # ── Palette & helpers ─────────────────────────────────────────────────────
    NAVY   = "0F1D2E"
    BLUE   = "1668B8"
    LTBLUE = "E6F1FB"
    GRAY   = "F5F7FA"
    BDGRAY = "E0E6ED"
    WHITE  = "FFFFFF"

    def hdr_font(size=11, bold=True, color=WHITE):
        return Font(name="Calibri", size=size, bold=bold, color=color)
    def body_font(size=10, bold=False, color="0F1D2E"):
        return Font(name="Calibri", size=size, bold=bold, color=color)
    def center_align(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def left_align(wrap=False):
        return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

    def set_col_width(ws, col_letter, width):
        ws.column_dimensions[col_letter].width = width

    def write_header_row(ws, row_num, headers, bg=BLUE, fg=WHITE, height=22):
        ws.row_dimensions[row_num].height = height
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=ci, value=h)
            c.font = hdr_font(color=fg)
            c.fill = _cell_fill(bg)
            c.alignment = center_align()
            c.border = _thin_border()

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1: Executive Summary
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_view.showGridLines = False

    # Title block
    ws1.merge_cells("A1:H1")
    c = ws1["A1"]
    c.value = "TerriSense — Territory Optimization Report"
    c.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    c.fill = _cell_fill(NAVY)
    c.alignment = center_align()
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:H2")
    c = ws1["A2"]
    c.value = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  K={k}  |  Tolerance ±{tolerance_pct}%  |  Floor={hard_floor}  |  Range {min_cap}–{max_cap}"
    c.font = Font(name="Calibri", size=10, color="7D8A98")
    c.fill = _cell_fill("F5F7FA")
    c.alignment = center_align()
    ws1.row_dimensions[2].height = 18

    # Scorecard tiles (row 4)
    green_n  = sum(1 for s in stats.values() if s['Status'] == 'Green')
    yellow_n = sum(1 for s in stats.values() if s['Status'] == 'Yellow')
    orange_n = sum(1 for s in stats.values() if s['Status'] == 'Orange')
    red_n    = sum(1 for s in stats.values() if s['Status'] == 'Red')
    total_n  = len(stats)

    tiles = [
        ("Total Territories", total_n,  BLUE,    WHITE),
        ("✓ Optimal",          green_n,  "1AA15A", WHITE),
        ("⚠ Slightly Low",    yellow_n, "D18A17", WHITE),
        ("⚠ Slightly High",   orange_n, "E07020", WHITE),
        ("✗ Below Floor",      red_n,    "D13B3B", WHITE),
    ]
    ws1.row_dimensions[4].height = 14
    ws1.row_dimensions[5].height = 28
    ws1.row_dimensions[6].height = 20
    ws1.row_dimensions[7].height = 14

    for ci, (label, val, bg, fg) in enumerate(tiles, 1):
        col = get_column_letter(ci)
        ws1.merge_cells(f"{col}5:{col}6")
        # Value
        c = ws1.cell(row=5, column=ci, value=val)
        c.font = Font(name="Calibri", size=20, bold=True, color=fg)
        c.fill = _cell_fill(bg)
        c.alignment = center_align()
        # Label
        cl = ws1.cell(row=7, column=ci, value=label)
        cl.font = Font(name="Calibri", size=9, bold=False, color="4A5B6E")
        cl.alignment = center_align()

    # Set tile columns wide
    for ci in range(1, 6):
        ws1.column_dimensions[get_column_letter(ci)].width = 18

    # AI Executive Summary
    ws1.row_dimensions[9].height = 16
    ws1.merge_cells("A9:H9")
    c = ws1["A9"]
    c.value = "AI EXECUTIVE SUMMARY"
    c.font = Font(name="Calibri", size=11, bold=True, color=BLUE)
    c.alignment = left_align()

    ws1.row_dimensions[10].height = 70
    ws1.merge_cells("A10:H10")
    c = ws1["A10"]
    c.value = ai.get("executive_summary", "")
    c.font = body_font(size=10)
    c.fill = _cell_fill(LTBLUE)
    c.alignment = left_align(wrap=True)
    c.border = _thin_border()

    # Strategic Recommendation
    ws1.row_dimensions[12].height = 16
    ws1.merge_cells("A12:H12")
    c = ws1["A12"]
    c.value = "STRATEGIC RECOMMENDATION"
    c.font = Font(name="Calibri", size=11, bold=True, color=BLUE)
    c.alignment = left_align()

    ws1.row_dimensions[13].height = 70
    ws1.merge_cells("A13:H13")
    c = ws1["A13"]
    c.value = ai.get("strategic_recommendation", "")
    c.font = body_font(size=10)
    c.fill = _cell_fill("F1F4F8")
    c.alignment = left_align(wrap=True)
    c.border = _thin_border()

    # Outlier suggestions table
    suggestions = ai.get("outlier_suggestions", [])
    if suggestions:
        ws1.row_dimensions[15].height = 16
        ws1.merge_cells("A15:H15")
        c = ws1["A15"]
        c.value = "OUTLIER TERRITORY SUGGESTIONS"
        c.font = Font(name="Calibri", size=11, bold=True, color=BLUE)
        c.alignment = left_align()

        write_header_row(ws1, 16, ["Territory", "Issue", "AI Suggestion"], bg=NAVY)
        ws1.column_dimensions["A"].width = 14
        ws1.column_dimensions["B"].width = 35
        ws1.merge_cells("C16:H16")
        ws1.column_dimensions["C"].width = 55

        for ri, sug in enumerate(suggestions, 17):
            ws1.row_dimensions[ri].height = 32
            cells = [
                (1, sug.get("territory", "")),
                (2, sug.get("issue", "")),
                (3, sug.get("suggestion", "")),
            ]
            for ci, val in cells:
                c = ws1.cell(row=ri, column=ci, value=val)
                c.font = body_font()
                c.border = _thin_border()
                c.alignment = left_align(wrap=True)
                if ci == 1:
                    c.font = Font(name="Calibri", size=10, bold=True, color=BLUE)
                c.fill = _cell_fill(GRAY if ri % 2 == 0 else WHITE)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2: Territory Summary (styled, conditional)
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Territory Summary")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    c = ws2["A1"]
    c.value = "Territory Summary"
    c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    c.fill = _cell_fill(NAVY)
    c.alignment = center_align()
    ws2.row_dimensions[1].height = 28

    hdrs2 = ["Territory", "Status", "Index", "vs Target", "Active ZIPs", "Total ZIPs", "Diameter (mi)", "Centroid Lat", "Centroid Lon"]
    write_header_row(ws2, 2, hdrs2)
    col_widths2 = [14, 14, 12, 14, 14, 12, 16, 14, 14]
    for ci, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    sorted_stats = sorted(stats.items(), key=lambda x: x[1]['Weight'], reverse=True)
    for ri, (t_id, s) in enumerate(sorted_stats, 3):
        ws2.row_dimensions[ri].height = 18
        status = s['Status']
        bg_hex, fg_hex = STATUS_FILLS.get(status, ("FFFFFF", "0F1D2E"))
        vs_target = s['Weight'] - 1000

        row_vals = [
            f"T{t_id}", status, s['Weight'],
            f"+{vs_target}" if vs_target >= 0 else str(vs_target),
            s['NonZeroZips'], s['ZipCount'], s['Diameter'],
            s.get('Centroid_Lat', ''), s.get('Centroid_Lon', '')
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.border = _thin_border()
            c.alignment = center_align()
            if ci in (1, 2):
                c.fill = _cell_fill(bg_hex)
                c.font = Font(name="Calibri", size=10, bold=(ci == 1), color=fg_hex)
            else:
                c.font = body_font()
                c.fill = _cell_fill(GRAY if ri % 2 == 0 else WHITE)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 3: ZIP Assignments
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("ZIP Assignments")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:F1")
    c = ws3["A1"]
    c.value = "ZIP Code Territory Assignments"
    c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    c.fill = _cell_fill(NAVY)
    c.alignment = center_align()
    ws3.row_dimensions[1].height = 28

    hdrs3 = ["ZIP Code", "Territory", "Index", "HCP Count", "Latitude", "Longitude"]
    write_header_row(ws3, 2, hdrs3)
    col_widths3 = [14, 14, 12, 14, 14, 14]
    for ci, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    active_only = df_res[df_res['final_weight'] > 0].sort_values('Territory_ID')
    for ri, (_, row) in enumerate(active_only.iterrows(), 3):
        ws3.row_dimensions[ri].height = 16
        vals = [
            row['clean_zip'], f"T{int(row['Territory_ID'])}",
            int(row['final_weight']), int(row.get('rep_count', 0)),
            round(float(row['latitude']), 4), round(float(row['longitude']), 4),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.font = body_font()
            c.border = _thin_border()
            c.alignment = center_align()
            c.fill = _cell_fill(GRAY if ri % 2 == 0 else WHITE)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 4: Outlier Deep-Dive
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Outlier Analysis")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:G1")
    c = ws4["A1"]
    c.value = "Outlier Territory Analysis"
    c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    c.fill = _cell_fill("D13B3B")
    c.alignment = center_align()
    ws4.row_dimensions[1].height = 28

    outliers = {t: s for t, s in stats.items() if s['Status'] != 'Green'}
    if outliers:
        hdrs4 = ["Territory", "Status", "Index", "Gap to Target", "Active ZIPs", "Diameter (mi)", "Action Needed"]
        write_header_row(ws4, 2, hdrs4, bg="99231F")
        col_widths4 = [14, 14, 12, 16, 14, 16, 45]
        for ci, w in enumerate(col_widths4, 1):
            ws4.column_dimensions[get_column_letter(ci)].width = w

        # Map AI suggestions to territory IDs
        ai_sugs = {s['territory']: s['suggestion']
                   for s in ai.get('outlier_suggestions', [])}

        for ri, (t_id, s) in enumerate(
                sorted(outliers.items(), key=lambda x: x[1]['Weight']), 3):
            ws4.row_dimensions[ri].height = 32
            status = s['Status']
            bg_hex, fg_hex = STATUS_FILLS.get(status, ("FFFFFF", "0F1D2E"))
            gap = s['Weight'] - 1000
            action = ai_sugs.get(f"T{t_id}", s['Message'])

            vals = [
                f"T{t_id}", status, s['Weight'],
                f"+{gap}" if gap >= 0 else str(gap),
                s['NonZeroZips'], s['Diameter'], action
            ]
            for ci, val in enumerate(vals, 1):
                c = ws4.cell(row=ri, column=ci, value=val)
                c.border = _thin_border()
                c.alignment = left_align(wrap=True) if ci == 7 else center_align()
                if ci in (1, 2):
                    c.fill = _cell_fill(bg_hex)
                    c.font = Font(name="Calibri", size=10, bold=(ci==1), color=fg_hex)
                else:
                    c.font = body_font()
                    c.fill = _cell_fill(GRAY if ri % 2 == 0 else WHITE)
    else:
        ws4.merge_cells("A3:G3")
        c = ws4["A3"]
        c.value = "🎉 All territories are optimal — no outliers detected!"
        c.font = Font(name="Calibri", size=12, bold=True, color="0E6E3C")
        c.fill = _cell_fill("E3F5EC")
        c.alignment = center_align()
        ws4.row_dimensions[3].height = 30

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ---------------------------------------------------------------------------
# HCP-level analysis helpers (unchanged)
# ---------------------------------------------------------------------------
def compute_hcp_composite(df, column_config):
    df = df.copy()
    numeric_cols = []
    for col in column_config:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            numeric_cols.append(col)
    if not numeric_cols:
        raise ValueError(
            f"None of the specified columns found. "
            f"Specified: {list(column_config.keys())}. File has: {df.columns.tolist()}"
        )
    df['_composite'] = 0.0
    for col, pct in column_config.items():
        if col in numeric_cols:
            col_max = df[col].max()
            norm = df[col] / col_max if col_max > 0 else 0.0
            df['_composite'] += (float(pct) / 100.0) * norm
    return df


def assign_deciles(df):
    df = df.copy()
    total = df['_composite'].sum()
    if total <= 0:
        df['_decile'] = 0
        return df
    df = df.sort_values('_composite', ascending=False).reset_index(drop=True)
    df['_cumshare'] = df['_composite'].cumsum() / total
    prev = df['_cumshare'].shift(1).fillna(0)
    df['_decile'] = prev.apply(lambda x: max(0, min(10, 10 - int(x * 10))))
    return df


def assign_segments(df, segment_config):
    df = df.copy()
    df['_segment'] = 'Unassigned'
    df['_calls_per_hcp'] = 0
    df['_targeted'] = False
    for seg in segment_config:
        if not seg.get('target', False):
            continue
        lo = min(int(seg['start_decile']), int(seg['end_decile']))
        hi = max(int(seg['start_decile']), int(seg['end_decile']))
        mask = (df['_decile'] >= lo) & (df['_decile'] <= hi)
        df.loc[mask, '_segment'] = seg['segment']
        df.loc[mask, '_calls_per_hcp'] = float(seg.get('calls_per_hcp', 0))
        df.loc[mask, '_targeted'] = True
    return df


def compute_calls_required(df):
    return int((df['_calls_per_hcp'] * df['_targeted']).sum())


def build_decile_table(df):
    rows = []
    for d in range(10, -1, -1):
        subset = df[df['_decile'] == d]
        calls = int(subset['_calls_per_hcp'].max()) if len(subset) > 0 else 0
        rows.append({"decile": f"D{d}", "hcp_count": len(subset), "calls_per_hcp": calls})
    return rows


def detect_metric_columns(df):
    non_metric = {
        'customer_id','customerid','id','customer_name','customername','name',
        'city','state','zip','zipcode','zip_code','postalcode','postal_code',
        'county','country','zip_population','zippopulation','population',
    }
    derived = ['normalize','composite','index','decile','segment','score','unnamed']
    metrics = []
    for col in df.columns:
        if not isinstance(col, str):
            continue
        key = col.lower().replace('_','').replace(' ','').replace(':','')
        if any(key == nm.replace('_','') for nm in non_metric):
            continue
        if any(kw in col.lower() for kw in derived):
            continue
        if key.startswith('unnamed'):
            continue
        try:
            if pd.to_numeric(df[col], errors='coerce').notna().mean() >= 0.5:
                metrics.append(col)
        except Exception:
            pass
    return metrics


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
@app.get("/health")
def health():
    groq_key = os.environ.get("GROQ_API_KEY", "")
    return {
        "status": "ok",
        "groq_key_set": bool(groq_key),
        "groq_key_preview": (groq_key[:8] + "…" if groq_key else "NOT SET"),
    }


@app.post("/detect_columns")
async def detect_columns(file: UploadFile = File(...)):
    try:
        raw = await file.read()
        df  = read_file_smartly(raw, file.filename)
        metrics     = detect_metric_columns(df)
        all_columns = df.columns.tolist()

        # Build a 5-row preview for the UI
        preview_df   = df.head(5).fillna("").astype(str)
        preview_rows = preview_df.values.tolist()

        return {
            "columns":     metrics,          # auto-detected metric cols
            "all_columns": all_columns,      # every column in the file
            "preview": {
                "headers": all_columns,
                "rows":    preview_rows,
            },
        }
    except Exception as e:
        return {"error": str(e)}


@app.post("/analyze")
async def analyze(
    file: UploadFile = File(...),
    column_config: str = Form(...),
    segment_config: str = Form(...),
):
    try:
        config   = json.loads(column_config)
        segments = json.loads(segment_config)
        content  = await file.read()
        df = read_file_smartly(content, file.filename)
        df = compute_hcp_composite(df, config)
        df = assign_deciles(df)
        df = assign_segments(df, segments)
        calls_required = compute_calls_required(df)
        decile_table   = build_decile_table(df)
        seg_summary = []
        for seg in segments:
            if not seg.get('target', False):
                continue
            name = seg['segment']
            subset = df[df['_segment'] == name]
            seg_summary.append({
                "segment": name,
                "hcp_count": len(subset),
                "calls_per_hcp": seg.get('calls_per_hcp', 0),
                "total_calls": int(subset['_calls_per_hcp'].sum()),
                "start_decile": seg['start_decile'],
                "end_decile": seg['end_decile'],
            })
        return {
            "calls_required": calls_required,
            "decile_table": decile_table,
            "segment_summary": seg_summary,
            "total_hcps": len(df),
            "targeted_hcps": int(df['_targeted'].sum()),
        }
    except Exception as e:
        print(f"ERROR /analyze: {e}")
        return {"error": str(e)}


@app.post("/optimize_map")
async def optimize_map(
    file: UploadFile = File(...),
    num_clusters: int = Form(...),
    column_config: str = Form(...),
    tolerance_pct: int = Form(15),
):
    """Generate territory map only. Independent from Excel endpoint."""
    try:
        config     = json.loads(column_config)
        raw_bytes  = await file.read()
        raw_df     = read_file_smartly(raw_bytes, file.filename)
        base_df    = preprocess_data(raw_df, config)

        hard_floor = 650.0
        min_cap    = round(1000 * (1 - tolerance_pct / 100))
        max_cap    = round(1000 * (1 + tolerance_pct / 100))
        k          = int(num_clusters)

        df_res, stats_res, k_rec = run_scenario(base_df, k, min_cap, max_cap, hard_floor)
        gc.collect()

        html_str = generate_map_html(
            df_res, stats_res, k,
            f"TerriSense — K={k} | Tolerance ±{tolerance_pct}%"
        )
        gc.collect()

        # Generate AI analysis here where stats are available
        ai_result = generate_ai_inference(
            stats_res, k, tolerance_pct, int(hard_floor), min_cap, max_cap
        )
        gc.collect()

        headers = {
            'X-Optimal-K':   str(k_rec['optimal_k']),
            'X-K-Min':       str(k_rec['k_min']),
            'X-K-Max':       str(k_rec['k_max']),
            'X-AI-Analysis': json.dumps(ai_result),
            'Access-Control-Expose-Headers': 'X-Optimal-K, X-K-Min, X-K-Max, X-AI-Analysis',
        }
        return Response(content=html_str, media_type="text/html", headers=headers)
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"ERROR /optimize_map: {e}")
        return {"error": str(e)}


@app.post("/optimize_excel")
async def optimize_excel(
    file: UploadFile = File(...),
    num_clusters: int = Form(...),
    column_config: str = Form(...),
    tolerance_pct: int = Form(15),
):
    """Generate Excel report only. Independent from Map endpoint."""
    try:
        config     = json.loads(column_config)
        raw_bytes  = await file.read()
        raw_df     = read_file_smartly(raw_bytes, file.filename)
        base_df    = preprocess_data(raw_df, config)

        hard_floor = 650
        min_cap    = round(1000 * (1 - tolerance_pct / 100))
        max_cap    = round(1000 * (1 + tolerance_pct / 100))
        k          = int(num_clusters)

        df_res, stats_res, k_rec = run_scenario(base_df, k, min_cap, max_cap, hard_floor)
        gc.collect()

        ai         = generate_ai_inference(stats_res, k, tolerance_pct, hard_floor, min_cap, max_cap)
        xlsx_bytes = build_excel(df_res, stats_res, k_rec, k, tolerance_pct,
                                 hard_floor, min_cap, max_cap, ai)
        gc.collect()

        headers = {
            'Content-Disposition':         'attachment; filename="territory_analysis.xlsx"',
            'X-Optimal-K':                 str(k_rec['optimal_k']),
            'Access-Control-Expose-Headers': 'X-Optimal-K',
        }
        return Response(
            content=xlsx_bytes,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers,
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"ERROR /optimize_excel: {e}")
        return {"error": str(e)}


@app.post("/ai_inference")
async def ai_inference(
    stats_json: str = Form(...),
    k: int = Form(...),
    tolerance_pct: int = Form(15),
    hard_floor: int = Form(650),
    min_cap: int = Form(850),
    max_cap: int = Form(1150),
):
    """Standalone endpoint for on-platform AI inference panel."""
    try:
        stats = json.loads(stats_json)
        # Convert string keys back to int
        stats = {int(k_): v for k_, v in stats.items()}
        result = generate_ai_inference(stats, k, tolerance_pct, hard_floor, min_cap, max_cap)
        return result
    except Exception as e:
        print(f"ERROR /ai_inference: {e}")
        return {"error": str(e)}
